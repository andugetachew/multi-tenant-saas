import pandas as pd
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from projects.models import Project, Task
from accounts.models import User


class ExcelReportGenerator:
    def __init__(self, user):
        self.user = user
        self.org = user.organization

    def generate_comprehensive_report(self):
        # Create workbook
        wb = Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        # Sheet 1: Projects Summary
        ws_projects = wb.create_sheet("Projects Summary")
        projects = Project.objects.filter(organization=self.org)

        # Headers
        headers = ["ID", "Project Name", "Description", "Tasks", "Created Date"]
        ws_projects.append(headers)

        for project in projects:
            ws_projects.append(
                [
                    project.id,
                    project.name,
                    project.description[:50],
                    project.tasks.count(),
                    project.created_at.strftime("%Y-%m-%d"),
                ]
            )

        # Add chart
        chart = BarChart()
        chart.title = "Projects by Month"
        data = Reference(
            ws_projects, min_col=4, min_row=1, max_col=4, max_row=len(projects) + 1
        )
        chart.add_data(data, titles_from_data=True)
        ws_projects.add_chart(chart, "G2")

        # Sheet 2: Tasks Summary
        ws_tasks = wb.create_sheet("Tasks Summary")
        tasks = Task.objects.filter(project__organization=self.org)

        ws_tasks.append(["Task Title", "Project", "Status", "Priority", "Due Date"])
        for task in tasks:
            ws_tasks.append(
                [
                    task.title,
                    task.project.name,
                    task.status,
                    task.priority,
                    task.due_date.strftime("%Y-%m-%d") if task.due_date else "",
                ]
            )

        # Sheet 3: Team Performance
        ws_team = wb.create_sheet("Team Performance")
        ws_team.append(["User", "Tasks Created", "Comments", "Productivity Score"])

        for user in User.objects.filter(organization=self.org):
            tasks_count = Task.objects.filter(created_by=user).count()
            comments_count = user.comment_set.count()
            score = tasks_count * 10 + comments_count

            ws_team.append([user.email, tasks_count, comments_count, score])

        # Save response
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="saas_report.xlsx"'
        wb.save(response)

        return response
